import os
import sys
import calendar
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Asegurar importación correcta de scripts
DIR_ACTUAL = os.path.dirname(os.path.abspath(__file__))
DIR_PROYECTO = os.path.dirname(DIR_ACTUAL)
if DIR_PROYECTO not in sys.path:
    sys.path.insert(0, DIR_PROYECTO)

# Importar funciones del motor de corrección
from scripts.motor_correccion_multiple import (
    cargar_archivo_csv,
    corregir_causa_motivo_40,
    corregir_codtecnologia_sin_ceros,
    corregir_ium_por_cum,
    reubicar_fechas_mes_objetivo,
    corregir_finalidades_por_diagnostico,
    corregir_sexo_por_reglas,
    corregir_pais_origen,
    corregir_condicion_destino_egreso,
    corregir_diagnosticos_duplicados,
    corregir_num_consultas_prenatal,
    log_correcciones,
    determinar_mes_y_anio_objetivo
)

DIR_ESTRUCTURA_FEV = r"C:\GESTION_FEV_EAPB\FEV_EAPB"
DIR_REPORTES_EXCEL = r"C:\GESTION_FEV_EAPB\REPORTES_EXCEL"
RUTA_SALIDA_EXCEL = os.path.join(DIR_REPORTES_EXCEL, "Reporte_Consolidado_Mes_vs_EAPB_Coloreado.xlsx")

MESES_ORDEN = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
EAPBS_LISTA = ["EMSSANAR", "NUEVA EPS", "ASMET SALUD", "COOSALUD", "ENTIDAD PROMOTORA (S.O.S)"]

def procesar_y_contar_factura(ruta_csv):
    """Carga el CSV, ejecuta el pipeline de auditoría/corrección y retorna los conteos exactos."""
    global log_correcciones
    log_correcciones.clear()

    df = cargar_archivo_csv(ruta_csv)
    if df.empty:
        return 0, 0, 0

    df = df.fillna("")
    reg_totales = len(df)
    
    num_usuarios = 0
    if 'numDocumentoIdentificacion' in df.columns:
        num_usuarios = df['numDocumentoIdentificacion'].nunique()

    # Ejecutar pipeline de correcciones para auditoría
    anio_calc, mes_calc, _ = determinar_mes_y_anio_objetivo(df)
    df = corregir_causa_motivo_40(df)
    df = corregir_codtecnologia_sin_ceros(df)
    df = corregir_ium_por_cum(df)
    df = reubicar_fechas_mes_objetivo(df, anio_calc, mes_calc)
    df = corregir_finalidades_por_diagnostico(df)
    df = corregir_sexo_por_reglas(df)
    df = corregir_pais_origen(df)
    df = corregir_condicion_destino_egreso(df)
    df = corregir_diagnosticos_duplicados(df)
    df = corregir_num_consultas_prenatal(df)

    total_inconsistencias = len(log_correcciones)
    return reg_totales, num_usuarios, total_inconsistencias


def recolectar_datos():
    """Recorre la estructura C:\\GESTION_FEV_EAPB\\FEV_EAPB y recolecta métricas de todas las facturas CSV."""
    registros = []
    print("=" * 80)
    print("ANALIZANDO FACTURAS Y GENERANDO MÉTRICAS DESDE C:\\GESTION_FEV_EAPB")
    print("=" * 80)

    if not os.path.exists(DIR_ESTRUCTURA_FEV):
        raise FileNotFoundError(f"No existe la ruta: {DIR_ESTRUCTURA_FEV}")

    for root, _, files in os.walk(DIR_ESTRUCTURA_FEV):
        for f in files:
            if f.lower().endswith('.csv'):
                ruta = os.path.join(root, f)
                rel_path = os.path.relpath(ruta, DIR_ESTRUCTURA_FEV)
                parts = rel_path.split(os.sep)

                eapb = parts[0]
                mes_folder = parts[1] if len(parts) > 1 else ''
                mes_nombre = mes_folder.split('.')[-1] if '.' in mes_folder else mes_folder

                print(f"  [+] Procesando ({eapb} - {mes_nombre}): {f} ...")
                reg_totales, num_usuarios, inconsistencias = procesar_y_contar_factura(ruta)

                registros.append({
                    'Mes': mes_nombre,
                    'EAPB': eapb,
                    'Factura': os.path.splitext(f)[0],
                    'Registros Totales': reg_totales,
                    'Usuarios': num_usuarios,
                    'Inconsistencias Corregidas': inconsistencias,
                    'Archivo Fuente': f
                })

    df_detalle = pd.DataFrame(registros)
    
    # Ordenar detalle por mes y EAPB
    df_detalle['Mes_Num'] = df_detalle['Mes'].apply(lambda x: MESES_ORDEN.index(x) if x in MESES_ORDEN else 99)
    df_detalle = df_detalle.sort_values(by=['Mes_Num', 'EAPB', 'Factura']).drop(columns=['Mes_Num'])

    return df_detalle


def generar_excel_estilizado(df_detalle):
    """Crea el libro Excel con formato coloreado, 2 pestañas y totales alineados."""
    os.makedirs(DIR_REPORTES_EXCEL, exist_ok=True)
    
    wb = openpyxl.Workbook()
    # Eliminar hoja default
    wb.remove(wb.active)

    # ------------------------------------------------------------------------
    # HOJA 1: CONSOLIDADO MES VS EAPB
    # ------------------------------------------------------------------------
    ws1 = wb.create_sheet(title="Consolidado Mes vs EAPB")
    ws1.views.sheetView[0].showGridLines = True

    # Definir paleta de colores elegantes
    color_header_bg = "1F4E78"     # Azul Oscuro
    color_header_fg = "FFFFFF"     # Blanco
    color_eapb_bg = "2F5597"       # Azul Medio
    color_total_bg = "D9E1F2"      # Azul Claro
    color_total_gen = "8EA9DB"     # Azul Acentuado
    color_zebra = "F2F2F2"         # Gris Suave

    font_header = Font(name="Calibri", size=11, bold=True, color=color_header_fg)
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_normal = Font(name="Calibri", size=11)
    
    fill_header_top = PatternFill(start_color=color_header_bg, end_color=color_header_bg, fill_type="solid")
    fill_eapb = PatternFill(start_color=color_eapb_bg, end_color=color_eapb_bg, fill_type="solid")
    fill_total_mes = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_total_gen = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")
    fill_zebra = PatternFill(start_color=color_zebra, end_color=color_zebra, fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Estructura de cabeceras en ws1
    ws1.cell(row=1, column=1, value="MES").fill = fill_header_top
    ws1.cell(row=1, column=1).font = font_header
    ws1.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws1.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    col_idx = 2
    for eapb in EAPBS_LISTA:
        ws1.cell(row=1, column=col_idx, value=eapb).fill = fill_header_top
        ws1.cell(row=1, column=col_idx).font = font_header
        ws1.cell(row=1, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
        ws1.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+2)

        ws1.cell(row=2, column=col_idx, value="Registros").fill = fill_eapb
        ws1.cell(row=2, column=col_idx+1, value="Usuarios").fill = fill_eapb
        ws1.cell(row=2, column=col_idx+2, value="Inconsist. Corregidas").fill = fill_eapb

        for c in range(col_idx, col_idx+3):
            ws1.cell(row=2, column=c).font = font_header
            ws1.cell(row=2, column=c).alignment = Alignment(horizontal="center", vertical="center")

        col_idx += 3

    # Columna TOTAL MES
    ws1.cell(row=1, column=col_idx, value="TOTAL MES").fill = fill_header_top
    ws1.cell(row=1, column=col_idx).font = font_header
    ws1.cell(row=1, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
    ws1.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+2)

    ws1.cell(row=2, column=col_idx, value="Registros").fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    ws1.cell(row=2, column=col_idx+1, value="Usuarios").fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    ws1.cell(row=2, column=col_idx+2, value="Tot. Inconsist. Corregida").fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")

    for c in range(col_idx, col_idx+3):
        ws1.cell(row=2, column=c).font = font_header
        ws1.cell(row=2, column=c).alignment = Alignment(horizontal="center", vertical="center")

    # Poblar filas por Mes
    meses_presentes = [m for m in MESES_ORDEN if m in df_detalle['Mes'].unique()]
    row_idx = 3

    for idx_m, mes in enumerate(meses_presentes):
        is_even = (idx_m % 2 == 0)
        row_fill = fill_zebra if is_even else PatternFill(fill_type=None)

        ws1.cell(row=row_idx, column=1, value=mes).font = font_bold
        ws1.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws1.cell(row=row_idx, column=1).fill = row_fill
        ws1.cell(row=row_idx, column=1).border = thin_border

        c_offset = 2
        mes_total_reg = 0
        mes_total_usr = 0
        mes_total_inc = 0

        for eapb in EAPBS_LISTA:
            sub_df = df_detalle[(df_detalle['Mes'] == mes) & (df_detalle['EAPB'] == eapb)]
            reg = int(sub_df['Registros Totales'].sum()) if not sub_df.empty else 0
            usr = int(sub_df['Usuarios'].sum()) if not sub_df.empty else 0
            inc = int(sub_df['Inconsistencias Corregidas'].sum()) if not sub_df.empty else 0

            mes_total_reg += reg
            mes_total_usr += usr
            mes_total_inc += inc

            for c_pos, val in enumerate([reg, usr, inc]):
                cell = ws1.cell(row=row_idx, column=c_offset + c_pos, value=val)
                cell.font = font_normal
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.fill = row_fill
                cell.border = thin_border

            c_offset += 3

        # Totales del Mes
        for c_pos, val in enumerate([mes_total_reg, mes_total_usr, mes_total_inc]):
            cell = ws1.cell(row=row_idx, column=c_offset + c_pos, value=val)
            cell.font = font_bold
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.fill = fill_total_mes
            cell.border = thin_border

        row_idx += 1

    # Fila TOTAL GENERAL
    ws1.cell(row=row_idx, column=1, value="TOTAL GENERAL").font = font_bold
    ws1.cell(row=row_idx, column=1).fill = fill_total_gen
    ws1.cell(row=row_idx, column=1).border = thin_border

    c_offset = 2
    gen_total_reg = 0
    gen_total_usr = 0
    gen_total_inc = 0

    for eapb in EAPBS_LISTA:
        sub_df = df_detalle[df_detalle['EAPB'] == eapb]
        reg = int(sub_df['Registros Totales'].sum()) if not sub_df.empty else 0
        usr = int(sub_df['Usuarios'].sum()) if not sub_df.empty else 0
        inc = int(sub_df['Inconsistencias Corregidas'].sum()) if not sub_df.empty else 0

        gen_total_reg += reg
        gen_total_usr += usr
        gen_total_inc += inc

        for c_pos, val in enumerate([reg, usr, inc]):
            cell = ws1.cell(row=row_idx, column=c_offset + c_pos, value=val)
            cell.font = font_bold
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.fill = fill_total_gen
            cell.border = thin_border

        c_offset += 3

    # Total General Final
    for c_pos, val in enumerate([gen_total_reg, gen_total_usr, gen_total_inc]):
        cell = ws1.cell(row=row_idx, column=c_offset + c_pos, value=val)
        cell.font = font_bold
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.fill = fill_total_gen
        cell.border = thin_border

    # Adjust Column Widths en ws1
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 14)

    # ------------------------------------------------------------------------
    # HOJA 2: DETALLE FACTURAS FEV
    # ------------------------------------------------------------------------
    ws2 = wb.create_sheet(title="Detalle Facturas FEV")
    ws2.views.sheetView[0].showGridLines = True

    headers_ws2 = ["Mes", "EAPB", "Factura", "Registros Totales", "Usuarios", "Inconsistencias Corregidas", "Archivo Fuente"]
    ws2.append(headers_ws2)

    for col_num in range(1, len(headers_ws2) + 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.fill = fill_header_top
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, row in df_detalle.iterrows():
        ws2.append([
            row['Mes'],
            row['EAPB'],
            row['Factura'],
            row['Registros Totales'],
            row['Usuarios'],
            row['Inconsistencias Corregidas'],
            row['Archivo Fuente']
        ])
        curr_row = ws2.max_row
        is_even = (curr_row % 2 == 0)
        row_fill = fill_zebra if is_even else PatternFill(fill_type=None)

        for col_num in range(1, len(headers_ws2) + 1):
            cell = ws2.cell(row=curr_row, column=col_num)
            cell.font = font_normal
            cell.fill = row_fill
            cell.border = thin_border
            if col_num in [4, 5, 6]:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Adjust Column Widths en ws2
    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 4, 15)

    # Guardar Libro con protección de archivo abierto
    try:
        wb.save(RUTA_SALIDA_EXCEL)
        print("=" * 80)
        print(f"REPORTE CONSOLIDADO EXCEL GENERADO CON ÉXITO EN:\n{RUTA_SALIDA_EXCEL}")
        print("=" * 80)
        return RUTA_SALIDA_EXCEL
    except PermissionError:
        ruta_salida_alt = os.path.join(DIR_REPORTES_EXCEL, "Reporte_Consolidado_Mes_vs_EAPB_Coloreado_NUEVO.xlsx")
        wb.save(ruta_salida_alt)
        print("=" * 80)
        print(f"[!] Advertencia: {RUTA_SALIDA_EXCEL} está abierto en Excel.")
        print(f"REPORTE CONSOLIDADO GENERADO EN ARCHIVO ALTERNATIVO:\n{ruta_salida_alt}")
        print("=" * 80)
        return ruta_salida_alt


if __name__ == '__main__':
    df_det = recolectar_datos()
    generar_excel_estilizado(df_det)
